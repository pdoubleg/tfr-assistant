import json
import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.api.routers.chat import router as chat_router
from app.core.config import Settings, get_settings
from app.models.chat_state import TFRChatState
from app.services.chat_artifacts import ChatArtifactStore
from app.services.output_bundles import OutputBundleError, OutputBundleService


def _settings(tmp_path: Path) -> Settings:
    return Settings(
        data_dir=tmp_path / "data",
        chat_artifacts_dir=tmp_path / "data" / "chat_artifacts",
    )


def test_report_bundle_renders_html_manifest_spec_and_workbook(tmp_path) -> None:
    settings = _settings(tmp_path)
    state = TFRChatState(artifact_session_id="report-test")
    store = ChatArtifactStore(settings)
    dataset_handle = store.save_dataset(
        state,
        columns=["driver", "count", "formula_text"],
        rows=[["Missing doc", 4, "=unsafe"], ["Late file", 2, "ok"]],
        label="Top drivers",
        source="test",
    ).handle
    chart_handle = store.save_plotly_chart(
        state,
        figure={
            "data": [
                {
                    "type": "bar",
                    "x": ["Missing doc", "Late file"],
                    "y": [4, 2],
                    "marker": {"color": "#636efa"},
                }
            ],
            "layout": {"title": {"text": "Default Plotly color"}},
        },
        label="Top drivers chart",
        source=dataset_handle,
    ).handle
    service = OutputBundleService(settings)

    report_handle = service.create_report_bundle(
        state,
        title="Audit Exception Summary",
        subtitle="Financial claim review results",
    )
    service.add_report_block(
        state,
        report_handle,
        {"type": "markdown", "title": "Summary", "markdown": "Findings are **ready**."},
    )
    service.add_report_block(
        state,
        report_handle,
        {
            "type": "table",
            "dataset_handle": dataset_handle,
            "title": "Top Drivers",
            "columns": ["driver", "count", "formula_text"],
            "max_rows": 1,
        },
    )
    service.add_report_block(
        state,
        report_handle,
        {"type": "chart", "chart_handle": chart_handle, "title": "Top Drivers Chart"},
    )

    payload = service.render_report_bundle(state, report_handle)

    assert payload["handle"] == "rpt_1"
    assert [file["role"] for file in payload["files"]] == ["html", "spec", "data", "manifest"]
    bundle_dir = settings.chat_artifacts_dir / state.artifact_session_id / report_handle
    assert (bundle_dir / "report.html").exists()
    assert (bundle_dir / "report-spec.json").exists()
    assert (bundle_dir / "manifest.json").exists()
    assert (bundle_dir / "data.xlsx").exists()

    html = (bundle_dir / "report.html").read_text(encoding="utf-8")
    assert "#FFD000" in html
    assert "#1A1446" in html
    assert "Showing 1 of 2 row(s)" in html
    assert 'src="https://cdn.plot.ly' not in html
    assert '"marker":{"color":"#1A1446"}' in html

    workbook = openpyxl.load_workbook(bundle_dir / "data.xlsx", data_only=False)
    assert {"README", "manifest", "Top drivers"}.issubset(set(workbook.sheetnames))
    sheet = workbook["Top drivers"]
    assert sheet["A2"].value == "Missing doc"
    assert sheet["C2"].value == "'=unsafe"

    manifest = json.loads((bundle_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_handles"] == [dataset_handle, chart_handle]
    assert manifest["files"][-1]["role"] == "manifest"
    assert state.handles[-1].kind == "report_bundle"


def test_deck_bundle_renders_valid_pptx_without_charts(tmp_path) -> None:
    if shutil.which("node") is None:
        pytest.skip("Node.js is required for PPTX rendering.")
    if not (Path(__file__).resolve().parents[1] / "node_modules" / "pptxgenjs").exists():
        pytest.skip("pptxgenjs is not installed.")

    settings = _settings(tmp_path)
    state = TFRChatState(artifact_session_id="deck-test")
    store = ChatArtifactStore(settings)
    dataset_handle = store.save_dataset(
        state,
        columns=["driver", "count"],
        rows=[["Missing doc", 4], ["Late file", 2], ["Other", 1]],
        label="Top drivers",
        source="test",
    ).handle
    service = OutputBundleService(settings)

    deck_handle = service.create_deck_bundle(state, title="Audit Briefing")
    service.add_deck_slide(state, deck_handle, {"type": "title", "title": "Audit Briefing"})
    service.add_deck_slide(
        state,
        deck_handle,
        {"type": "table", "title": "Top Drivers", "dataset_handle": dataset_handle, "max_rows": 2},
    )

    payload = service.render_deck_bundle(state, deck_handle)

    assert payload["handle"] == "deck_1"
    assert [file["role"] for file in payload["files"]] == ["spec", "pptx", "data", "manifest"]
    bundle_dir = settings.chat_artifacts_dir / state.artifact_session_id / deck_handle
    pptx_path = bundle_dir / "deck.pptx"
    assert zipfile.is_zipfile(pptx_path)
    with zipfile.ZipFile(pptx_path) as archive:
        assert "ppt/slides/slide1.xml" in archive.namelist()
        assert "ppt/slides/slide2.xml" in archive.namelist()

    workbook = openpyxl.load_workbook(bundle_dir / "data.xlsx", data_only=True)
    assert workbook["Top drivers"].max_row == 4
    spec = json.loads((bundle_dir / "deck-spec.json").read_text(encoding="utf-8"))
    assert spec["slides"][1]["rendered_rows"] == 2
    assert spec["slides"][1]["row_count"] == 3
    assert state.handles[-1].kind == "deck_bundle"


def test_output_bundle_rejects_invalid_handles_and_custom_elements(tmp_path) -> None:
    settings = _settings(tmp_path)
    state = TFRChatState(artifact_session_id="invalid-test")
    service = OutputBundleService(settings)

    report_handle = service.create_report_bundle(state, title="Invalid Report")

    with pytest.raises(OutputBundleError, match="Expected dataset handle"):
        service.add_report_block(
            state,
            report_handle,
            {"type": "table", "dataset_handle": "fig_1", "title": "Wrong"},
        )

    deck_handle = service.create_deck_bundle(state, title="Invalid Deck")
    with pytest.raises(OutputBundleError, match="must fit inside"):
        service.add_deck_slide(
            state,
            deck_handle,
            {
                "type": "custom",
                "title": "Bad",
                "elements": [
                    {"kind": "text", "text": "Too wide", "box": {"x": 13, "y": 1, "w": 1, "h": 1}}
                ],
            },
        )


def test_chat_artifact_file_endpoint_serves_manifest_backed_files(tmp_path) -> None:
    settings = _settings(tmp_path)
    state = TFRChatState(artifact_session_id="api-test")
    service = OutputBundleService(settings)
    report_handle = service.create_report_bundle(state, title="API Report")
    service.add_report_block(
        state,
        report_handle,
        {"type": "markdown", "title": "Summary", "markdown": "Ready."},
    )
    service.render_report_bundle(state, report_handle, include_workbook=False)

    app = FastAPI()
    app.dependency_overrides[get_settings] = lambda: settings
    app.include_router(chat_router, prefix="/api/chat")
    client = TestClient(app)

    ok = client.get(f"/api/chat/artifacts/{state.artifact_session_id}/{report_handle}/files/html")
    assert ok.status_code == 200
    assert "text/html" in ok.headers["content-type"]
    assert "API Report" in ok.text

    missing = client.get(
        f"/api/chat/artifacts/{state.artifact_session_id}/{report_handle}/files/pptx"
    )
    assert missing.status_code == 404

    traversal = client.get(f"/api/chat/artifacts/../{report_handle}/files/html")
    assert traversal.status_code == 404
