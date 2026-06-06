"""Rendered report and deck bundle services for Monty output generation."""

from __future__ import annotations

import html
import json
import re
import subprocess
from collections.abc import Iterable
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Literal

import plotly.graph_objects as go
import plotly.io as pio
from jinja2 import Environment, select_autoescape
from markupsafe import Markup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from plotly.offline import get_plotlyjs

from app.core.config import Settings
from app.models.chat_state import TFRChatState
from app.services.chat_artifacts import (
    ChatArtifactStore,
    DatasetArtifact,
    OutputBundleArtifact,
    OutputBundleFile,
    PlotlyChartArtifact,
)

BundleKind = Literal["report_bundle", "deck_bundle"]

LIBERTY_PALETTE = {
    "yellow": "FFD000",
    "blue": "1A1446",
    "teal": "78E1E1",
    "dark_teal": "037B86",
    "atmospheric_gray": "F5F5F5",
    "white": "FFFFFF",
    "dark_gray": "343741",
    "black": "000000",
}
LIBERTY_CHART_COLORWAY = [
    f"#{LIBERTY_PALETTE['blue']}",
    f"#{LIBERTY_PALETTE['dark_teal']}",
    f"#{LIBERTY_PALETTE['teal']}",
    f"#{LIBERTY_PALETTE['yellow']}",
    f"#{LIBERTY_PALETTE['dark_gray']}",
]
LIBERTY_CONTINUOUS_COLORSCALE = [
    [0.0, f"#{LIBERTY_PALETTE['blue']}"],
    [0.5, f"#{LIBERTY_PALETTE['dark_teal']}"],
    [1.0, f"#{LIBERTY_PALETTE['teal']}"],
]

MAX_SLIDE_WIDTH = 13.333
MAX_SLIDE_HEIGHT = 7.5
CUSTOM_ELEMENT_KINDS = {"text", "shape", "callout", "metric", "table", "chart"}


class OutputBundleError(ValueError):
    """Raised when an output bundle cannot be created or rendered."""


class OutputBundleService:
    """Create, update, and render durable report/deck output bundles."""

    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.store = ChatArtifactStore(settings)

    def create_report_bundle(
        self,
        state: TFRChatState,
        *,
        title: str,
        subtitle: str = "",
        audience: str = "audit",
        theme: str = "liberty_professional",
    ) -> str:
        bundle = self.store.save_output_bundle(
            state,
            kind="report_bundle",
            title=_required_text(title, "title"),
            subtitle=subtitle.strip(),
            audience=audience.strip() or "audit",
            theme=_theme(theme),
        )
        return bundle.handle

    def create_deck_bundle(
        self,
        state: TFRChatState,
        *,
        title: str,
        subtitle: str = "",
        audience: str = "executive",
        theme: str = "liberty_professional",
        layout: str = "wide",
    ) -> str:
        if layout != "wide":
            raise OutputBundleError("Only layout='wide' is supported in v1.")
        bundle = self.store.save_output_bundle(
            state,
            kind="deck_bundle",
            title=_required_text(title, "title"),
            subtitle=subtitle.strip(),
            audience=audience.strip() or "executive",
            theme=_theme(theme),
            layout=layout,
        )
        return bundle.handle

    def add_report_block(
        self,
        state: TFRChatState,
        report_handle: str,
        block: dict[str, Any],
    ) -> str:
        bundle = self._load_bundle(state, report_handle, kind="report_bundle")
        block = self._validate_report_block(state, block)
        block["id"] = f"block_{len(bundle.blocks) + 1}"
        bundle.blocks.append(block)
        bundle.source_handles = _merge_source_handles(
            bundle.source_handles,
            _source_handles_from_block(block),
        )
        self.store.write_output_bundle(state, bundle)
        return bundle.handle

    def add_deck_slide(
        self,
        state: TFRChatState,
        deck_handle: str,
        slide: dict[str, Any],
    ) -> str:
        bundle = self._load_bundle(state, deck_handle, kind="deck_bundle")
        slide = self._validate_deck_slide(state, slide)
        slide["id"] = f"slide_{len(bundle.slides) + 1}"
        bundle.slides.append(slide)
        bundle.source_handles = _merge_source_handles(
            bundle.source_handles,
            _source_handles_from_slide(slide),
        )
        self.store.write_output_bundle(state, bundle)
        return bundle.handle

    def render_report_bundle(
        self,
        state: TFRChatState,
        report_handle: str,
        *,
        output_name: str = "",
        include_workbook: bool = True,
    ) -> dict[str, Any]:
        bundle = self._load_bundle(state, report_handle, kind="report_bundle")
        bundle_dir = self._prepare_bundle_dir(state, bundle, output_name=output_name)
        warnings: list[str] = []
        files: list[OutputBundleFile] = []

        body_blocks = [self._report_block_view(state, block) for block in bundle.blocks]
        html_content = self._render_report_html(bundle, body_blocks)
        html_path = bundle_dir / "report.html"
        html_path.write_text(html_content, encoding="utf-8", newline="")
        files.append(
            _file_record(
                "html",
                "report.html",
                "text/html",
                f"{_slug(bundle.title)}.html",
                "HTML report",
                inline=True,
            )
        )

        spec_path = bundle_dir / "report-spec.json"
        spec_path.write_text(
            json.dumps(_bundle_spec(bundle), ensure_ascii=False, indent=2),
            encoding="utf-8",
            newline="",
        )
        files.append(
            _file_record(
                "spec",
                "report-spec.json",
                "application/json",
                f"{_slug(bundle.title)}-spec.json",
                "Report spec",
            )
        )

        if include_workbook:
            workbook = self._write_data_workbook(state, bundle, bundle_dir, warnings=warnings)
            if workbook is not None:
                files.append(workbook)

        return self._finalize_rendered_bundle(
            state,
            bundle,
            bundle_dir,
            files=files,
            warnings=warnings,
        )

    def render_deck_bundle(
        self,
        state: TFRChatState,
        deck_handle: str,
        *,
        output_name: str = "",
        include_workbook: bool = True,
    ) -> dict[str, Any]:
        bundle = self._load_bundle(state, deck_handle, kind="deck_bundle")
        bundle_dir = self._prepare_bundle_dir(state, bundle, output_name=output_name)
        assets_dir = bundle_dir / "assets"
        assets_dir.mkdir(parents=True, exist_ok=True)
        warnings: list[str] = []
        files: list[OutputBundleFile] = []

        deck_spec = self._deck_spec(state, bundle, assets_dir)
        spec_path = bundle_dir / "deck-spec.json"
        spec_path.write_text(
            json.dumps(deck_spec, ensure_ascii=False, indent=2),
            encoding="utf-8",
            newline="",
        )
        files.append(
            _file_record(
                "spec",
                "deck-spec.json",
                "application/json",
                f"{_slug(bundle.title)}-deck-spec.json",
                "Deck spec",
            )
        )

        pptx_path = bundle_dir / "deck.pptx"
        self._render_pptx(spec_path, pptx_path)
        files.append(
            _file_record(
                "pptx",
                "deck.pptx",
                "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                f"{_slug(bundle.title)}.pptx",
                "PowerPoint deck",
            )
        )

        if include_workbook:
            workbook = self._write_data_workbook(state, bundle, bundle_dir, warnings=warnings)
            if workbook is not None:
                files.append(workbook)

        return self._finalize_rendered_bundle(
            state,
            bundle,
            bundle_dir,
            files=files,
            warnings=warnings,
        )

    def bundle_payload(self, state: TFRChatState, bundle_handle: str) -> dict[str, Any]:
        bundle = self.store.load_output_bundle(state, bundle_handle)
        return _bundle_card_payload(state.artifact_session_id, bundle)

    def _load_bundle(
        self,
        state: TFRChatState,
        handle: str,
        *,
        kind: BundleKind,
    ) -> OutputBundleArtifact:
        bundle = self.store.load_output_bundle(state, handle)
        if bundle.kind != kind:
            raise OutputBundleError(f"Handle {handle!r} is not a {kind}.")
        return bundle

    def _validate_report_block(
        self,
        state: TFRChatState,
        block: dict[str, Any],
    ) -> dict[str, Any]:
        block_type = str(block.get("type") or "")
        normalized = dict(block)
        if block_type == "kpis":
            normalized["metrics"] = _normalize_metrics(block.get("metrics") or [])
            return normalized
        if block_type == "markdown":
            _required_text(str(block.get("title") or ""), "title")
            return normalized
        if block_type == "chart":
            chart_handle = str(block.get("chart_handle") or "")
            self._load_chart(state, chart_handle)
            normalized["chart_handle"] = chart_handle
            return normalized
        if block_type == "table":
            dataset_handle = str(block.get("dataset_handle") or "")
            dataset = self._load_dataset(state, dataset_handle)
            if block.get("columns") is not None:
                _selected_columns(dataset, block.get("columns"))
            if int(block.get("max_rows") or 50) < 1:
                raise OutputBundleError("max_rows must be at least 1.")
            normalized["dataset_handle"] = dataset_handle
            return normalized
        raise OutputBundleError(f"Unsupported report block type: {block_type}")

    def _validate_deck_slide(
        self,
        state: TFRChatState,
        slide: dict[str, Any],
    ) -> dict[str, Any]:
        slide_type = str(slide.get("type") or "")
        normalized = dict(slide)
        if slide_type == "title":
            return normalized
        if slide_type == "metrics":
            normalized["metrics"] = _normalize_metrics(slide.get("metrics") or [])
            return normalized
        if slide_type == "findings":
            normalized["findings"] = _normalize_findings(slide.get("findings") or [])
            return normalized
        if slide_type == "chart":
            chart_handle = str(slide.get("chart_handle") or "")
            self._load_chart(state, chart_handle)
            normalized["chart_handle"] = chart_handle
            return normalized
        if slide_type == "table":
            dataset_handle = str(slide.get("dataset_handle") or "")
            dataset = self._load_dataset(state, dataset_handle)
            if slide.get("columns") is not None:
                _selected_columns(dataset, slide.get("columns"))
            if int(slide.get("max_rows") or 12) < 1:
                raise OutputBundleError("max_rows must be at least 1.")
            normalized["dataset_handle"] = dataset_handle
            return normalized
        if slide_type == "custom":
            normalized["elements"] = validate_custom_elements(slide.get("elements") or [])
            return normalized
        raise OutputBundleError(f"Unsupported deck slide type: {slide_type}")

    def _prepare_bundle_dir(
        self,
        state: TFRChatState,
        bundle: OutputBundleArtifact,
        *,
        output_name: str,
    ) -> Path:
        bundle_dir = self.store.output_bundle_dir(state, bundle.handle)
        (bundle_dir / "assets").mkdir(parents=True, exist_ok=True)
        name = output_name.strip() or f"{_slug(bundle.title)}-{_timestamp_slug()}"
        (bundle_dir / "output-name.txt").write_text(name, encoding="utf-8", newline="")
        return bundle_dir

    def _report_block_view(self, state: TFRChatState, block: dict[str, Any]) -> dict[str, Any]:
        block_type = str(block.get("type") or "")
        if block_type == "kpis":
            return {
                "type": "kpis",
                "title": block.get("title") or "",
                "metrics": _normalize_metrics(block.get("metrics") or []),
            }
        if block_type == "markdown":
            return {
                "type": "markdown",
                "title": block.get("title") or "",
                "html": Markup(_markdown_to_html(str(block.get("markdown") or ""))),
            }
        if block_type == "chart":
            chart = self._load_chart(state, str(block.get("chart_handle") or ""))
            return {
                "type": "chart",
                "title": block.get("title") or chart.label,
                "caption": block.get("caption") or "",
                "width": block.get("width") or "full",
                "chart_html": Markup(self._plotly_html(chart)),
            }
        if block_type == "table":
            dataset = self._load_dataset(state, str(block.get("dataset_handle") or ""))
            table = _dataset_table(
                dataset,
                columns=block.get("columns"),
                max_rows=int(block.get("max_rows") or 50),
            )
            return {
                "type": "table",
                "title": block.get("title") or dataset.label,
                "caption": block.get("caption") or "",
                **table,
            }
        raise OutputBundleError(f"Unsupported report block type: {block_type}")

    def _render_report_html(
        self,
        bundle: OutputBundleArtifact,
        blocks: list[dict[str, Any]],
    ) -> str:
        env = Environment(autoescape=select_autoescape(["html", "xml"]))
        template = env.from_string(REPORT_TEMPLATE)
        rendered_at = datetime.now(UTC)
        plotly_js = ""
        if any(block.get("type") == "chart" for block in blocks):
            plotly_js = Markup(get_plotlyjs())
        return template.render(
            palette={key: f"#{value}" for key, value in LIBERTY_PALETTE.items()},
            chart_colorway=LIBERTY_CHART_COLORWAY,
            plotly_js=plotly_js,
            bundle=bundle,
            blocks=blocks,
            rendered_at=rendered_at.isoformat(timespec="seconds"),
            rendered_at_label=rendered_at.strftime("%b %d, %Y %H:%M UTC"),
        )

    def _deck_spec(
        self,
        state: TFRChatState,
        bundle: OutputBundleArtifact,
        assets_dir: Path,
    ) -> dict[str, Any]:
        slides = [self._deck_slide_spec(state, slide, assets_dir) for slide in bundle.slides]
        if not slides:
            slides.append(
                {
                    "type": "title",
                    "title": bundle.title,
                    "subtitle": bundle.subtitle,
                    "kicker": "Generated output",
                    "notes": "",
                }
            )
        return {
            "schema_version": 1,
            "title": bundle.title,
            "subtitle": bundle.subtitle,
            "audience": bundle.audience,
            "theme": bundle.theme,
            "layout": bundle.layout or "wide",
            "palette": LIBERTY_PALETTE,
            "created_at": datetime.now(UTC).isoformat(timespec="seconds"),
            "slides": slides,
        }

    def _deck_slide_spec(
        self,
        state: TFRChatState,
        slide: dict[str, Any],
        assets_dir: Path,
    ) -> dict[str, Any]:
        slide_type = str(slide.get("type") or "")
        notes = str(slide.get("notes") or "")
        if slide_type == "title":
            return {
                "type": "title",
                "title": slide.get("title") or "",
                "subtitle": slide.get("subtitle") or "",
                "kicker": slide.get("kicker") or "",
                "notes": notes,
            }
        if slide_type == "metrics":
            return {
                "type": "metrics",
                "title": slide.get("title") or "",
                "metrics": _normalize_metrics(slide.get("metrics") or []),
                "notes": notes,
            }
        if slide_type == "findings":
            return {
                "type": "findings",
                "title": slide.get("title") or "",
                "findings": _normalize_findings(slide.get("findings") or []),
                "notes": notes,
            }
        if slide_type == "chart":
            chart_handle = str(slide.get("chart_handle") or "")
            image_path = self._export_chart_image(state, chart_handle, assets_dir)
            return {
                "type": "chart",
                "title": slide.get("title") or "",
                "caption": slide.get("caption") or "",
                "imagePath": str(image_path),
                "notes": notes,
            }
        if slide_type == "table":
            dataset = self._load_dataset(state, str(slide.get("dataset_handle") or ""))
            return {
                "type": "table",
                "title": slide.get("title") or dataset.label,
                **_dataset_table(
                    dataset,
                    columns=slide.get("columns"),
                    max_rows=int(slide.get("max_rows") or 12),
                ),
                "notes": notes,
            }
        if slide_type == "custom":
            return {
                "type": "custom",
                "title": slide.get("title") or "",
                "elements": [
                    self._custom_element_spec(state, element, assets_dir)
                    for element in slide.get("elements") or []
                ],
                "notes": notes,
            }
        raise OutputBundleError(f"Unsupported deck slide type: {slide_type}")

    def _custom_element_spec(
        self,
        state: TFRChatState,
        element: dict[str, Any],
        assets_dir: Path,
    ) -> dict[str, Any]:
        kind = str(element.get("kind") or "")
        if kind not in CUSTOM_ELEMENT_KINDS:
            raise OutputBundleError(f"Unsupported custom slide element kind: {kind}")
        box = _validated_box(element.get("box") or {}, argument_name="element.box")
        resolved = dict(element)
        resolved["box"] = box
        if kind == "table" and element.get("dataset_handle"):
            dataset = self._load_dataset(state, str(element.get("dataset_handle")))
            resolved.update(
                _dataset_table(
                    dataset,
                    columns=element.get("columns"),
                    max_rows=int(element.get("max_rows") or 12),
                )
            )
        if kind == "chart" and element.get("chart_handle"):
            resolved["imagePath"] = str(
                self._export_chart_image(state, str(element.get("chart_handle")), assets_dir)
            )
        return resolved

    def _plotly_html(self, chart: PlotlyChartArtifact) -> str:
        figure = _themed_plotly_figure(chart.figure)
        return pio.to_html(
            figure,
            include_plotlyjs=False,
            full_html=False,
            config={"displaylogo": False, "responsive": True},
        )

    def _export_chart_image(
        self,
        state: TFRChatState,
        chart_handle: str,
        assets_dir: Path,
    ) -> Path:
        chart = self._load_chart(state, chart_handle)
        figure = _themed_plotly_figure(chart.figure)
        output = assets_dir / f"{chart.handle}.png"
        try:
            pio.write_image(figure, output, width=1280, height=720, scale=2)
        except Exception as exc:
            raise OutputBundleError(
                "Unable to export Plotly chart image for deck rendering. "
                "Install/configure Kaleido and Chrome/Chromium, or remove the chart slide."
            ) from exc
        return output

    def _write_data_workbook(
        self,
        state: TFRChatState,
        bundle: OutputBundleArtifact,
        bundle_dir: Path,
        *,
        warnings: list[str],
    ) -> OutputBundleFile | None:
        dataset_handles = self._dataset_handles_for_bundle(state, bundle)
        if not dataset_handles:
            warnings.append("No dataset handles were referenced, so no data workbook was created.")
            return None

        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_readme_sheet(workbook, bundle, dataset_handles)
        self._write_manifest_sheet(workbook, state, dataset_handles)
        used_sheet_names = {"README", "manifest"}
        for handle in dataset_handles:
            dataset = self._load_dataset(state, handle)
            sheet_name = _unique_sheet_name(
                _safe_sheet_name(dataset.label or handle),
                used_sheet_names,
            )
            used_sheet_names.add(sheet_name)
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(dataset.columns)
            for row in dataset.rows:
                sheet.append([_excel_safe_cell(value) for value in row])
            _style_data_sheet(sheet)
        output = bundle_dir / "data.xlsx"
        workbook.save(output)
        return _file_record(
            "data",
            "data.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{_slug(bundle.title)}-data.xlsx",
            "Data workbook",
        )

    def _write_readme_sheet(
        self,
        workbook: Workbook,
        bundle: OutputBundleArtifact,
        dataset_handles: list[str],
    ) -> None:
        sheet = workbook.create_sheet("README")
        rows = [
            ["Title", bundle.title],
            ["Subtitle", bundle.subtitle],
            ["Kind", bundle.kind],
            ["Handle", bundle.handle],
            ["Generated at", datetime.now(UTC).isoformat(timespec="seconds")],
            ["Dataset handles", ", ".join(dataset_handles)],
        ]
        for row in rows:
            sheet.append(row)
        _style_key_value_sheet(sheet)

    def _write_manifest_sheet(
        self,
        workbook: Workbook,
        state: TFRChatState,
        dataset_handles: list[str],
    ) -> None:
        sheet = workbook.create_sheet("manifest")
        sheet.append(["handle", "label", "rows", "columns", "source"])
        for handle in dataset_handles:
            dataset = self._load_dataset(state, handle)
            sheet.append(
                [
                    dataset.handle,
                    dataset.label,
                    dataset.row_count,
                    dataset.column_count,
                    dataset.source,
                ]
            )
        _style_data_sheet(sheet)

    def _dataset_handles_for_bundle(
        self,
        state: TFRChatState,
        bundle: OutputBundleArtifact,
    ) -> list[str]:
        handles: list[str] = []
        for handle in bundle.source_handles:
            if handle.startswith("ds_"):
                handles.append(handle)
            elif handle.startswith("fig_"):
                try:
                    chart = self.store.load_plotly_chart(state, handle)
                except Exception:
                    continue
                if chart.source.startswith("ds_"):
                    handles.append(chart.source)
        return list(dict.fromkeys(handles))

    def _load_dataset(self, state: TFRChatState, handle: str) -> DatasetArtifact:
        if not handle.startswith("ds_"):
            raise OutputBundleError(f"Expected dataset handle like 'ds_1', got {handle!r}.")
        try:
            return self.store.load_dataset(state, handle)
        except Exception as exc:
            raise OutputBundleError(f"Unknown dataset handle: {handle}") from exc

    def _load_chart(self, state: TFRChatState, handle: str) -> PlotlyChartArtifact:
        if not handle.startswith("fig_"):
            raise OutputBundleError(f"Expected chart handle like 'fig_1', got {handle!r}.")
        try:
            return self.store.load_plotly_chart(state, handle)
        except Exception as exc:
            raise OutputBundleError(f"Unknown chart handle: {handle}") from exc

    def _render_pptx(self, spec_path: Path, pptx_path: Path) -> None:
        backend_dir = Path(__file__).resolve().parents[2]
        script = backend_dir / "scripts" / "render_pptx.mjs"
        try:
            completed = subprocess.run(
                ["node", str(script), str(spec_path), str(pptx_path)],
                cwd=backend_dir,
                capture_output=True,
                text=True,
                timeout=90,
                check=False,
            )
        except FileNotFoundError as exc:
            raise OutputBundleError("Deck rendering requires Node.js on PATH.") from exc
        except subprocess.TimeoutExpired as exc:
            raise OutputBundleError("Deck rendering timed out.") from exc
        if completed.returncode != 0:
            message = (completed.stderr or completed.stdout or "Unknown PPTX render error").strip()
            raise OutputBundleError(f"Deck rendering failed: {message}")
        if not pptx_path.exists():
            raise OutputBundleError("Deck rendering completed but no PPTX file was created.")

    def _finalize_rendered_bundle(
        self,
        state: TFRChatState,
        bundle: OutputBundleArtifact,
        bundle_dir: Path,
        *,
        files: list[OutputBundleFile],
        warnings: list[str],
    ) -> dict[str, Any]:
        bundle.status = "rendered"
        bundle.files = files
        bundle.warnings = warnings
        manifest_record = _file_record(
            "manifest",
            "manifest.json",
            "application/json",
            f"{_slug(bundle.title)}-manifest.json",
            "Manifest",
        )
        bundle.files = [*bundle.files, manifest_record]
        manifest = {
            "schema_version": 1,
            "handle": bundle.handle,
            "kind": bundle.kind,
            "title": bundle.title,
            "subtitle": bundle.subtitle,
            "theme": bundle.theme,
            "status": bundle.status,
            "source_handles": bundle.source_handles,
            "files": [file.model_dump() for file in bundle.files],
            "warnings": warnings,
            "rendered_at": datetime.now(UTC).isoformat(timespec="seconds"),
        }
        (bundle_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
            newline="",
        )
        self.store.write_output_bundle(state, bundle)
        return _bundle_card_payload(state.artifact_session_id, bundle)


def validate_custom_elements(elements: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for index, element in enumerate(elements):
        if not isinstance(element, dict):
            raise OutputBundleError(f"Custom element at index {index} must be a dictionary.")
        kind = str(element.get("kind") or "")
        if kind not in CUSTOM_ELEMENT_KINDS:
            raise OutputBundleError(f"Unsupported custom slide element kind: {kind}")
        _validated_box(element.get("box") or {}, argument_name=f"elements[{index}].box")
        normalized.append(dict(element))
    return normalized


def _themed_plotly_figure(figure_payload: dict[str, Any]) -> go.Figure:
    figure = go.Figure(figure_payload)
    figure.update_layout(
        template="plotly_white",
        colorway=LIBERTY_CHART_COLORWAY,
        coloraxis={"colorscale": LIBERTY_CONTINUOUS_COLORSCALE},
        paper_bgcolor=f"#{LIBERTY_PALETTE['white']}",
        plot_bgcolor=f"#{LIBERTY_PALETTE['white']}",
        font={
            "family": "Aptos, Arial, sans-serif",
            "color": f"#{LIBERTY_PALETTE['dark_gray']}",
            "size": 13,
        },
        title_font={
            "color": f"#{LIBERTY_PALETTE['blue']}",
            "size": 18,
        },
        margin={"l": 56, "r": 32, "t": 64, "b": 56},
        hoverlabel={
            "bgcolor": f"#{LIBERTY_PALETTE['white']}",
            "bordercolor": f"#{LIBERTY_PALETTE['dark_teal']}",
            "font": {"color": f"#{LIBERTY_PALETTE['dark_gray']}"},
        },
    )
    figure.update_xaxes(gridcolor="#E7E7E7", linecolor=f"#{LIBERTY_PALETTE['dark_gray']}")
    figure.update_yaxes(gridcolor="#E7E7E7", linecolor=f"#{LIBERTY_PALETTE['dark_gray']}")
    _apply_liberty_trace_colors(figure)
    return figure


def _apply_liberty_trace_colors(figure: go.Figure) -> None:
    for index, trace in enumerate(figure.data):
        color = LIBERTY_CHART_COLORWAY[index % len(LIBERTY_CHART_COLORWAY)]
        trace_type = str(getattr(trace, "type", "") or "")

        if trace_type == "pie":
            point_count = _trace_point_count(trace)
            trace.update(marker={"colors": _cycled_palette(point_count)})
            continue

        marker = getattr(trace, "marker", None)
        if marker is not None:
            marker_color = getattr(marker, "color", None)
            if _is_color_sequence(marker_color):
                trace.update(marker={"color": _cycled_palette(len(marker_color))})
            elif not _is_numeric_sequence(marker_color):
                trace.update(marker={"color": color})

        if hasattr(trace, "line"):
            trace.update(line={"color": color})

        if hasattr(trace, "colorscale"):
            trace.update(colorscale=LIBERTY_CONTINUOUS_COLORSCALE)


def _trace_point_count(trace: Any) -> int:
    for attribute in ("labels", "values", "x", "y"):
        value = getattr(trace, attribute, None)
        if isinstance(value, list | tuple):
            return len(value)
    return len(LIBERTY_CHART_COLORWAY)


def _cycled_palette(count: int) -> list[str]:
    return [LIBERTY_CHART_COLORWAY[index % len(LIBERTY_CHART_COLORWAY)] for index in range(count)]


def _is_color_sequence(value: Any) -> bool:
    return isinstance(value, list | tuple) and any(isinstance(item, str) for item in value)


def _is_numeric_sequence(value: Any) -> bool:
    return isinstance(value, list | tuple) and all(isinstance(item, int | float) for item in value)


def _markdown_to_html(markdown: str) -> str:
    try:
        from markdown_it import MarkdownIt

        return MarkdownIt("commonmark", {"html": False}).enable("table").render(markdown)
    except Exception:
        escaped = html.escape(markdown)
        paragraphs = [part.strip() for part in re.split(r"\n\s*\n", escaped) if part.strip()]
        return "\n".join(f"<p>{part.replace(chr(10), '<br>')}</p>" for part in paragraphs)


def _dataset_table(
    dataset: DatasetArtifact,
    *,
    columns: Any,
    max_rows: int,
) -> dict[str, Any]:
    selected_columns = _selected_columns(dataset, columns)
    indexes = [dataset.columns.index(column) for column in selected_columns]
    row_limit = max(1, min(max_rows, 500))
    rows = [[row[index] for index in indexes] for row in dataset.rows[:row_limit]]
    return {
        "headers": selected_columns,
        "rows": rows,
        "row_count": dataset.row_count,
        "rendered_rows": len(rows),
    }


def _selected_columns(dataset: DatasetArtifact, columns: Any) -> list[str]:
    if columns is None:
        return list(dataset.columns)
    selected = [columns] if isinstance(columns, str) else list(columns)
    if not selected:
        raise OutputBundleError("columns cannot be an empty list.")
    missing = [column for column in selected if column not in dataset.columns]
    if missing:
        raise OutputBundleError(
            "Unknown dataset column(s): " + ", ".join(str(column) for column in missing)
        )
    return [str(column) for column in selected]


def _normalize_metrics(metrics: Any) -> list[dict[str, str]]:
    if not isinstance(metrics, list) or not metrics:
        raise OutputBundleError("metrics must be a non-empty list of dictionaries.")
    normalized: list[dict[str, str]] = []
    for index, metric in enumerate(metrics):
        if not isinstance(metric, dict):
            raise OutputBundleError(f"Metric at index {index} must be a dictionary.")
        label = _required_text(str(metric.get("label") or ""), f"metrics[{index}].label")
        value = _required_text(str(metric.get("value") or ""), f"metrics[{index}].value")
        normalized.append(
            {
                "label": label,
                "value": value,
                "detail": str(metric.get("detail") or ""),
            }
        )
    return normalized


def _normalize_findings(findings: Any) -> list[str]:
    if isinstance(findings, str):
        items = [line.strip("- ").strip() for line in findings.splitlines() if line.strip()]
    else:
        items = [str(item).strip() for item in list(findings or []) if str(item).strip()]
    if not items:
        raise OutputBundleError("findings must include at least one item.")
    return items


def _source_handles_from_block(block: dict[str, Any]) -> list[str]:
    handles = []
    for key in ("dataset_handle", "chart_handle"):
        value = block.get(key)
        if isinstance(value, str) and value:
            handles.append(value)
    return handles


def _source_handles_from_slide(slide: dict[str, Any]) -> list[str]:
    handles = _source_handles_from_block(slide)
    for element in slide.get("elements") or []:
        if isinstance(element, dict):
            handles.extend(_source_handles_from_block(element))
    return handles


def _merge_source_handles(existing: list[str], incoming: Iterable[str]) -> list[str]:
    return list(dict.fromkeys([*existing, *incoming]))


def _required_text(value: str, argument_name: str) -> str:
    text = value.strip()
    if not text:
        raise OutputBundleError(f"{argument_name} cannot be empty.")
    return text


def _theme(value: str) -> str:
    theme = value.strip() or "liberty_professional"
    if theme != "liberty_professional":
        raise OutputBundleError("Only theme='liberty_professional' is supported in v1.")
    return theme


def _validated_box(value: dict[str, Any], *, argument_name: str) -> dict[str, float]:
    required = ["x", "y", "w", "h"]
    missing = [key for key in required if key not in value]
    if missing:
        raise OutputBundleError(f"{argument_name} is missing: {', '.join(missing)}")
    box = {key: float(value[key]) for key in required}
    if box["x"] < 0 or box["y"] < 0 or box["w"] <= 0 or box["h"] <= 0:
        raise OutputBundleError(f"{argument_name} must contain positive slide dimensions.")
    if box["x"] + box["w"] > MAX_SLIDE_WIDTH or box["y"] + box["h"] > MAX_SLIDE_HEIGHT:
        raise OutputBundleError(
            f"{argument_name} must fit inside {MAX_SLIDE_WIDTH} x {MAX_SLIDE_HEIGHT} inches."
        )
    return box


def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug[:80] or "output"


def _timestamp_slug() -> str:
    return datetime.now(UTC).strftime("%Y%m%d-%H%M%S")


def _file_record(
    role: str,
    path: str,
    media_type: str,
    filename: str,
    label: str,
    *,
    inline: bool = False,
) -> OutputBundleFile:
    return OutputBundleFile(
        role=role,
        path=path.replace("\\", "/"),
        media_type=media_type,
        filename=filename,
        label=label,
        inline=inline,
    )


def _bundle_spec(bundle: OutputBundleArtifact) -> dict[str, Any]:
    payload = bundle.model_dump()
    payload["files"] = []
    return payload


def _bundle_card_payload(session_id: str, bundle: OutputBundleArtifact) -> dict[str, Any]:
    noun = "report" if bundle.kind == "report_bundle" else "deck"
    return {
        "component": "a2ui.ArtifactBundleCard",
        "sessionId": session_id,
        "handle": bundle.handle,
        "kind": bundle.kind,
        "title": bundle.title,
        "subtitle": bundle.subtitle,
        "summary": f"Rendered {noun} bundle with {len(bundle.files)} file(s).",
        "files": [file.model_dump() for file in bundle.files],
        "warnings": bundle.warnings,
        "createdAt": bundle.updated_at,
    }


def _safe_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\\/?*\[\]:]", " ", value).strip()
    return cleaned[:31] or "Sheet"


def _unique_sheet_name(base: str, used: set[str]) -> str:
    candidate = base[:31]
    index = 2
    while candidate in used:
        suffix = f" {index}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        index += 1
    return candidate


def _excel_safe_cell(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _style_data_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor=LIBERTY_PALETTE["blue"])
    header_font = Font(bold=True, color=LIBERTY_PALETTE["white"])
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(48, max(12, max_len + 2))


def _style_key_value_sheet(sheet: Any) -> None:
    for row in sheet.iter_rows():
        row[0].font = Font(bold=True, color=LIBERTY_PALETTE["blue"])
        row[0].alignment = Alignment(wrap_text=True)
        row[1].alignment = Alignment(wrap_text=True)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 80


REPORT_TEMPLATE = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ bundle.title }}</title>
  <style>
    :root {
      --liberty-yellow: {{ palette.yellow }};
      --liberty-blue: {{ palette.blue }};
      --liberty-teal: {{ palette.teal }};
      --liberty-dark-teal: {{ palette.dark_teal }};
      --liberty-gray: {{ palette.atmospheric_gray }};
      --liberty-white: {{ palette.white }};
      --liberty-dark-gray: {{ palette.dark_gray }};
      --liberty-black: {{ palette.black }};
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--liberty-gray);
      color: var(--liberty-dark-gray);
      font-family: Aptos, Arial, sans-serif;
      line-height: 1.48;
    }
    main {
      max-width: 1120px;
      margin: 0 auto;
      background: var(--liberty-white);
      min-height: 100vh;
      padding: 48px 56px 64px;
    }
    header {
      border-top: 12px solid var(--liberty-yellow);
      padding-top: 28px;
      margin-bottom: 36px;
    }
    h1, h2, h3 { color: var(--liberty-blue); margin: 0; letter-spacing: 0; }
    h1 { font-size: 34px; line-height: 1.08; }
    h2 { font-size: 20px; margin-bottom: 16px; }
    h3 { font-size: 15px; }
    .subtitle { color: var(--liberty-dark-teal); font-size: 16px; margin-top: 10px; }
    .meta { color: #656873; font-size: 12px; margin-top: 18px; }
    section { margin: 32px 0; break-inside: avoid; }
    section::before {
      content: "";
      display: block;
      width: 56px;
      height: 4px;
      background: var(--liberty-yellow);
      margin-bottom: 14px;
    }
    .kpi-grid {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 14px;
    }
    .kpi {
      border: 1px solid #E4E5E8;
      background: #FAFAFB;
      padding: 16px;
      border-radius: 4px;
    }
    .kpi-value {
      color: var(--liberty-blue);
      font-size: 28px;
      font-weight: 700;
    }
    .kpi-label { color: var(--liberty-dark-gray); font-weight: 700; margin-top: 2px; }
    .kpi-detail { color: #6B6E78; font-size: 12px; margin-top: 6px; }
    .markdown-body p { margin: 0 0 12px; }
    .markdown-body ul, .markdown-body ol { padding-left: 22px; }
    figure { margin: 0; }
    figcaption { color: #6B6E78; font-size: 12px; margin-top: 8px; }
    .table-wrap { overflow-x: auto; border: 1px solid #E4E5E8; }
    table { width: 100%; border-collapse: collapse; font-size: 12px; }
    th {
      background: var(--liberty-blue);
      color: var(--liberty-white);
      text-align: left;
      padding: 9px 10px;
      font-weight: 700;
    }
    td { border-top: 1px solid #ECEDEF; padding: 8px 10px; vertical-align: top; }
    tbody tr:nth-child(even) { background: #FAFAFB; }
    .table-note { color: #6B6E78; font-size: 11px; margin-top: 8px; }
    @media print {
      body { background: #fff; }
      main { max-width: none; padding: 24px; }
      a { color: inherit; text-decoration: none; }
    }
  </style>
  {% if plotly_js %}<script>{{ plotly_js }}</script>{% endif %}
</head>
<body>
  <main>
    <header>
      <h1>{{ bundle.title }}</h1>
      {% if bundle.subtitle %}<div class="subtitle">{{ bundle.subtitle }}</div>{% endif %}
      <div class="meta">Generated {{ rendered_at_label }} | {{ bundle.handle }}</div>
    </header>
    {% for block in blocks %}
      {% if block.type == "kpis" %}
        <section>
          {% if block.title %}<h2>{{ block.title }}</h2>{% endif %}
          <div class="kpi-grid">
            {% for metric in block.metrics %}
              <div class="kpi">
                <div class="kpi-value">{{ metric.value }}</div>
                <div class="kpi-label">{{ metric.label }}</div>
                {% if metric.detail %}<div class="kpi-detail">{{ metric.detail }}</div>{% endif %}
              </div>
            {% endfor %}
          </div>
        </section>
      {% elif block.type == "markdown" %}
        <section>
          {% if block.title %}<h2>{{ block.title }}</h2>{% endif %}
          <div class="markdown-body">{{ block.html }}</div>
        </section>
      {% elif block.type == "chart" %}
        <section>
          {% if block.title %}<h2>{{ block.title }}</h2>{% endif %}
          <figure>
            {{ block.chart_html }}
            {% if block.caption %}<figcaption>{{ block.caption }}</figcaption>{% endif %}
          </figure>
        </section>
      {% elif block.type == "table" %}
        <section>
          {% if block.title %}<h2>{{ block.title }}</h2>{% endif %}
          <div class="table-wrap">
            <table>
              <thead>
                <tr>
                  {% for header in block.headers %}<th>{{ header }}</th>{% endfor %}
                </tr>
              </thead>
              <tbody>
                {% for row in block.rows %}
                  <tr>
                    {% for cell in row %}<td>{{ "" if cell is none else cell }}</td>{% endfor %}
                  </tr>
                {% endfor %}
              </tbody>
            </table>
          </div>
          {% if block.caption %}<figcaption>{{ block.caption }}</figcaption>{% endif %}
          <div class="table-note">
            Showing {{ block.rendered_rows }} of {{ block.row_count }} row(s).
            Full data is in the workbook.
          </div>
        </section>
      {% endif %}
    {% endfor %}
  </main>
</body>
</html>
"""
